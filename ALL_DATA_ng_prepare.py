__version__ = "2026_0429_0841"
"""
ng_prepare.py - NG_Reasons 自動準備工具（程式1）
比對新版韌體的測試步驟 vs NG_Reasons 分頁，
自動插入尚未有說明的空白範本列，然後開啟 Excel 供填寫。
"""

import os
import re
import gc
import shutil
import logging
from datetime import datetime

import tkinter as tk
from tkinter import filedialog

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font

# ─── 設定區 ──────────────────────────────────────────────────────────────────
SEARCH_ROOT_DIR     = r"D:\yuanyu"            # 全域掃描根目錄（fallback 用）
YUANYU_ROOT         = SEARCH_ROOT_DIR         # 分頁名稱推斷專案目錄的基礎路徑
EXCLUDE_SHEET_NAMES = {"Index", "索引製作方法", "NG_Reasons"}
MIN_VERSION_DATE    = "20200101_0000"

# NG_Reasons: 每個 ERR 佔兩欄（意義 + 對照），最多支援幾個 ERR
MAX_ERR_COUNT = 4
# ─────────────────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - [%(levelname)s] - %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger()

VERSION_RE = re.compile(r"_(\d{8}_\d{4})")
STEP_RE    = re.compile(r"^#define\s+(\w+_STEP(?:_\d+)?)\s+(\d+)", re.MULTILINE)
EXCLUDE_KW = re.compile(r"last_data|last_time", re.IGNORECASE)

# NG_Reasons 固定欄（1-based）
COL_PRODUCT   = 1   # 產品
COL_VER_FROM  = 2   # 版本起
COL_STEP_CODE = 3   # 步驟代碼
COL_STEP_NAME = 4   # 步驟名稱
COL_DESC      = 5   # 不良描述
# ERR 欄從第 6 欄開始：(ERRn意義, ERRn對照) × MAX_ERR_COUNT
COL_ERR_START = 6

def err_col(n: int, is_map: bool) -> int:
    """ERR n（1-based）的欄號。is_map=False → 意義欄，is_map=True → 對照欄"""
    return COL_ERR_START + (n - 1) * 2 + (1 if is_map else 0)


# ═══════════════════════════════════════════════════════════════════════════════
# 選擇 Excel（檔案對話框）
# ═══════════════════════════════════════════════════════════════════════════════

def select_excel() -> str | None:
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    path = filedialog.askopenfilename(
        title="請選擇 Software Change Log Excel 檔案",
        filetypes=[("Excel 檔案", "*.xlsx *.xlsm"), ("所有檔案", "*.*")],
    )
    root.destroy()
    if not path:
        logger.warning("使用者取消選擇檔案")
        return None
    logger.info(f"已選擇: {os.path.basename(path)}")
    return path


# ═══════════════════════════════════════════════════════════════════════════════
# 建立路徑索引 + 填入 Excel F 欄（與 scanner.py 相同邏輯，ng_prepare 獨立可用）
# ═══════════════════════════════════════════════════════════════════════════════

def build_path_index(root_dir: str):
    logger.info(f"掃描 {root_dir} 建立路徑索引（請耐心等候）...")
    folder_index: dict[str, list[str]] = {}
    file_index:   dict[str, list[str]] = {}

    if not os.path.exists(root_dir):
        logger.error(f"路徑不存在: {root_dir}")
        return folder_index, file_index

    count = 0
    for dirpath, dirnames, filenames in os.walk(root_dir):
        count += 1
        if count % 1000 == 0:
            logger.info(f"  掃描中... 已遍歷 {count} 個子目錄")
        for d in dirnames:
            folder_index.setdefault(d, []).append(os.path.join(dirpath, d))
        for f in filenames:
            if f.lower().endswith(".hex"):
                # 無版本時間的 hex（開發中）直接忽略
                # 有版本時間但與所在資料夾路徑不一致的（複製帶入的舊 hex）也忽略
                m = VERSION_RE.search(f)
                if not m or m.group(1) not in dirpath:
                    continue
                file_index.setdefault(f, []).append(dirpath)

    logger.info(f"索引完成：資料夾 {len(folder_index)} 個，hex 檔 {len(file_index)} 個")
    return folder_index, file_index


def _best_path(candidates: list[str], sheet_name: str) -> str:
    """從多個候選路徑中，挑與 sheet_name 關鍵字最吻合的那條。"""
    if len(candidates) == 1:
        return candidates[0]
    keywords = re.findall(r"[A-Za-z0-9]+", sheet_name)
    def score(path: str) -> int:
        p = path.upper()
        return sum(1 for kw in keywords if kw.upper() in p)
    return max(candidates, key=score)


def get_real_cell(ws, cell):
    if isinstance(cell, MergedCell):
        for mr in ws.merged_cells.ranges:
            if cell.coordinate in mr:
                return ws.cell(mr.min_row, mr.min_col)
    return cell


def fill_excel_paths(ws, folder_idx: dict, file_idx: dict, sheet_name: str) -> int:
    """填寫 F 欄路徑，回傳更新筆數"""
    max_row = ws.max_row
    f1_val  = ws.cell(row=1, column=6).value

    if str(f1_val).strip() == "路徑":
        should_insert = False
    else:
        has_data = any(ws.cell(row=r, column=6).value for r in range(2, max_row + 1))
        should_insert = has_data
        if has_data:
            logger.info(f"  [{sheet_name}] F 欄含有其他資料，插入新欄位")

    if should_insert:
        ws.insert_cols(6)

    ws.cell(row=1, column=6).value = "路徑"

    updated = 0
    for row in range(2, max_row + 1):
        c_cell = ws.cell(row=row, column=3)
        if isinstance(c_cell, MergedCell) or not c_cell.value:
            continue
        key    = str(c_cell.value).strip()
        f_raw  = ws.cell(row=row, column=6)
        f_real = get_real_cell(ws, f_raw)

        if not key:
            f_real.value = None
            continue

        # 已有路徑則保留，不覆蓋
        cur = str(f_real.value or "").strip()
        if cur and cur.lower() != "none":
            continue

        candidates = (file_idx if ".hex" in key.lower() else folder_idx).get(key, [])
        path = _best_path(candidates, sheet_name) if candidates else None
        f_real.value = path
        if path:
            updated += 1

    return updated


# ═══════════════════════════════════════════════════════════════════════════════
# 分頁名稱 → 專案根目錄（精準掃描用）
# ═══════════════════════════════════════════════════════════════════════════════

def sheet_to_keys(sheet_name: str) -> list[str]:
    """
    從分頁名稱推導搜尋 key 清單（由精確到模糊）。
    '321-GL01A A' → ['321']
    '321A-FA01 A' → ['321A', '321']
    '816C-MGD1'   → ['816C', '816']
    '651電流偵測'  → ['651電流偵測', '651']
    'OEM-TY02'    → ['OEM']
    """
    keys: list[str] = []
    raw = sheet_name
    for sep in ("-", " ", "("):
        if sep in raw:
            raw = raw.split(sep)[0].strip()
            break
    keys.append(raw)
    # 數字+字母 → 補純數字 (321A→321, 816C→816)
    m = re.match(r'^(\d+)[A-Za-z]', raw)
    if m:
        keys.append(m.group(1))
    # ≥3位數字+非數字 → 補純數字 (651電流偵測→651)
    m2 = re.match(r'^(\d{3,})\D', raw)
    if m2 and m2.group(1) not in keys:
        keys.append(m2.group(1))
    return keys


def find_project_dir_by_name(sheet_name: str) -> str | None:
    """用分頁名稱 key 在 YUANYU_ROOT 第一層找對應資料夾。"""
    try:
        entries = [
            e for e in os.listdir(YUANYU_ROOT)
            if os.path.isdir(os.path.join(YUANYU_ROOT, e))
        ]
    except Exception:
        return None
    for key in sheet_to_keys(sheet_name):
        key_up = key.upper()
        for e in entries:
            if e.upper() == key_up:
                return os.path.join(YUANYU_ROOT, e)
        for e in entries:
            if e.upper().startswith(key_up):
                return os.path.join(YUANYU_ROOT, e)
        for e in entries:
            if key_up in e.upper():
                return os.path.join(YUANYU_ROOT, e)
    return None


def infer_project_dir_from_excel(ws) -> str | None:
    """從 F 欄現有路徑往上推，取 YUANYU_ROOT 下的第一層子目錄。"""
    yuanyu_depth = len(YUANYU_ROOT.replace("/", "\\").split("\\"))
    for row in range(2, ws.max_row + 1):
        c_cell = ws.cell(row=row, column=3)
        if isinstance(c_cell, MergedCell) or not c_cell.value:
            continue
        f_raw  = ws.cell(row=row, column=6)
        f_real = get_real_cell(ws, f_raw)
        path   = str(f_real.value or "").strip().replace("/", "\\")
        if not path or path.lower() in ("none", "疑似於筆記本電腦內"):
            continue
        if not os.path.exists(path):
            continue
        parts = path.split("\\")
        if len(parts) > yuanyu_depth:
            candidate = "\\".join(parts[:yuanyu_depth + 1])
            if os.path.isdir(candidate):
                return candidate
    return None


def get_project_dir(sheet_name: str, ws) -> str | None:
    """取得分頁專屬掃描根目錄：先名稱比對，再從 F 欄現有路徑推斷。"""
    return find_project_dir_by_name(sheet_name) or infer_project_dir_from_excel(ws)


def needs_path_filling(ws) -> bool:
    """是否有版本列的 F 欄尚未填寫路徑（有則需要掃描，否則可完全跳過）。"""
    for row in range(2, ws.max_row + 1):
        c_cell = ws.cell(row=row, column=3)
        if isinstance(c_cell, MergedCell) or not c_cell.value:
            continue
        if not VERSION_RE.search(str(c_cell.value)):
            continue
        f_raw  = ws.cell(row=row, column=6)
        f_real = get_real_cell(ws, f_raw)
        cur = str(f_real.value or "").strip()
        if not cur or cur.lower() == "none":
            return True
    return False


# ═══════════════════════════════════════════════════════════════════════════════
# 重建 Index 分頁
# ═══════════════════════════════════════════════════════════════════════════════

def rebuild_index_sheet(wb):
    """刪除舊 Index 分頁，重建所有工作表的超連結目錄"""
    if "Index" in wb.sheetnames:
        del wb["Index"]

    ws_idx = wb.create_sheet("Index", 0)
    ws_idx["A1"] = "工作表名稱"
    ws_idx["B1"] = "連結"

    for i, name in enumerate(wb.sheetnames):
        row = i + 2
        ws_idx.cell(row=row, column=1).value = name
        link_cell = ws_idx.cell(row=row, column=2)
        link_cell.value = "Go to Sheet"
        link_cell.hyperlink = f"#'{name}'!A1"
        try:
            link_cell.style = "Hyperlink"
        except Exception:
            link_cell.font = Font(color="0563C1", underline="single")

    ws_idx.column_dimensions["A"].width = 30
    ws_idx.column_dimensions["B"].width = 15
    logger.info("Index 分頁重建完成")


# ═══════════════════════════════════════════════════════════════════════════════
# 讀取韌體條目（從已填好 F 欄的 Excel）
# ═══════════════════════════════════════════════════════════════════════════════

def scan_firmware_entries(wb) -> list[dict]:
    entries: list[dict] = []
    for sheet_name in wb.sheetnames:
        if sheet_name in EXCLUDE_SHEET_NAMES or sheet_name == "Index":
            continue
        ws = wb[sheet_name]
        for row in range(2, ws.max_row + 1):
            c_cell = ws.cell(row=row, column=3)
            if isinstance(c_cell, MergedCell) or not c_cell.value:
                continue
            c_val = str(c_cell.value).strip()
            m = VERSION_RE.search(c_val)
            if not m:
                continue
            version_date = m.group(1)
            if version_date.replace("_", "") < MIN_VERSION_DATE.replace("_", ""):
                continue
            if "notYet" in c_val:
                continue

            f_cell = ws.cell(row=row, column=6)
            f_real = get_real_cell(ws, f_cell)
            folder_path = str(f_real.value).strip() if f_real.value else ""

            entries.append({
                "version_date": version_date,
                "folder_name":  c_val,
                "folder_path":  folder_path,
                "excel_sheet":  sheet_name.strip(),
            })
    return entries


# ═══════════════════════════════════════════════════════════════════════════════
# 解析 Define.h — 步驟
# ═══════════════════════════════════════════════════════════════════════════════

def parse_steps_from_define(folder_path: str) -> dict[str, str]:
    """回傳 {步驟數值字串: 步驟名稱}"""
    define_path = os.path.join(folder_path, "Head_file", "Define.h")
    if not os.path.exists(define_path):
        return {}
    try:
        with open(define_path, "r", encoding="utf-8", errors="ignore") as f:
            content = f.read()
    except Exception:
        return {}

    steps: dict[str, str] = {}
    for name, idx_str in STEP_RE.findall(content):
        if EXCLUDE_KW.search(name):
            continue
        steps[idx_str] = name
    return steps


# ═══════════════════════════════════════════════════════════════════════════════
# NG_Reasons 分頁管理
# ═══════════════════════════════════════════════════════════════════════════════

def build_ng_headers() -> list[str]:
    """產生 NG_Reasons 的標題列"""
    headers = ["產品", "版本起（空=全版本）", "步驟代碼", "步驟名稱", "不良描述"]
    for n in range(1, MAX_ERR_COUNT + 1):
        headers.append(f"ERR{n} 意義")
        headers.append(f"ERR{n} 對照（值=說明；分號分隔）")
    return headers


def ensure_ng_sheet(wb):
    """確保 NG_Reasons 分頁存在，不存在則建立並寫入標題列"""
    if "NG_Reasons" not in wb.sheetnames:
        ws_ng = wb.create_sheet("NG_Reasons")
        headers = build_ng_headers()
        for col, h in enumerate(headers, 1):
            ws_ng.cell(row=1, column=col).value = h

        # 欄寬設定
        col_widths = [14, 20, 10, 30, 30]
        for _ in range(MAX_ERR_COUNT):
            col_widths += [22, 38]
        for col, w in enumerate(col_widths, 1):
            ws_ng.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

        logger.info("已建立 NG_Reasons 分頁")
    return wb["NG_Reasons"]


def read_existing_ng(ws_ng) -> tuple[set, list]:
    """
    回傳:
      existing_keys : set of (product, ver_from, step_code)
      all_rows      : list of dict，含完整欄位（供前版繼承比對用）
    """
    existing_keys: set[tuple[str, str, str]] = set()
    all_rows: list[dict] = []

    for row in range(2, ws_ng.max_row + 1):
        product   = str(ws_ng.cell(row=row, column=COL_PRODUCT).value   or "").strip()
        ver_from  = str(ws_ng.cell(row=row, column=COL_VER_FROM).value  or "").strip()
        step_code = str(ws_ng.cell(row=row, column=COL_STEP_CODE).value or "").strip()
        desc      = str(ws_ng.cell(row=row, column=COL_DESC).value      or "").strip()
        if not product or not step_code:
            continue

        err_data: list[tuple[str, str]] = []
        for n in range(1, MAX_ERR_COUNT + 1):
            meaning = str(ws_ng.cell(row=row, column=err_col(n, False)).value or "").strip()
            map_raw = str(ws_ng.cell(row=row, column=err_col(n, True )).value or "").strip()
            err_data.append((meaning, map_raw))

        existing_keys.add((product, ver_from, step_code))
        all_rows.append({
            "product":   product,
            "ver_from":  ver_from,
            "step_code": step_code,
            "desc":      desc,
            "err_data":  err_data,
        })

    return existing_keys, all_rows


def find_prev_ng(all_rows: list, product: str, step_code: str, version_date: str) -> dict | None:
    """找同產品、同步驟、版本起 <= version_date 的最近一筆舊記錄"""
    candidates = [
        r for r in all_rows
        if r["product"] == product
        and r["step_code"] == step_code
        and (not r["ver_from"] or r["ver_from"] <= version_date)
    ]
    if not candidates:
        return None
    return max(candidates, key=lambda r: r["ver_from"])


# ═══════════════════════════════════════════════════════════════════════════════
# 主流程
# ═══════════════════════════════════════════════════════════════════════════════

def make_backup(excel_path: str) -> str | None:
    fdir  = os.path.dirname(excel_path)
    stem, ext = os.path.splitext(os.path.basename(excel_path))
    ts   = datetime.now().strftime("%Y%m%d_%H%M")
    # 把原始檔名中已有的 _YYYYMMDD_HHMM 替換為當下時間，再加 _AI
    new_stem = re.sub(r"_\d{8}_\d{4}(_AI)?$", "", stem) + f"_{ts}_AI"
    dest = os.path.join(fdir, new_stem + ext)
    try:
        shutil.copy2(excel_path, dest)
        logger.info(f"已建立副本: {os.path.basename(dest)}")
        return dest
    except Exception as e:
        logger.error(f"建立副本失敗: {e}")
        return None


def main():
    print("=" * 60)
    print(f"  NG_Reasons 自動準備工具 v{__version__}")
    print("=" * 60)

    # ── 選擇 Excel ────────────────────────────────────────────
    excel_path = select_excel()
    if not excel_path:
        input("未選擇檔案，按 Enter 鍵離開...")
        return

    # ── 建立副本 ──────────────────────────────────────────────
    backup_path = make_backup(excel_path)
    if not backup_path:
        input("備份失敗，程式中止。按 Enter 鍵離開...")
        return

    wb = None
    pending_items: list[dict] = []

    try:
        is_xlsm = backup_path.lower().endswith(".xlsm")
        wb = openpyxl.load_workbook(backup_path, keep_vba=is_xlsm)

        # ── 第一步：按分頁精準掃描，只填 F 欄空白列 ─────────────
        total_updated = 0
        for sheet_name in wb.sheetnames:
            if sheet_name in EXCLUDE_SHEET_NAMES or sheet_name == "Index":
                continue
            ws = wb[sheet_name]

            # 方案1：F 欄已全部填寫 → 略過，不啟動任何掃描
            if not needs_path_filling(ws):
                logger.info(f"  [{sheet_name}] F 欄已全部填寫，略過")
                continue

            # 方案2：找出該分頁的專屬掃描根目錄
            proj_dir = get_project_dir(sheet_name, ws)
            if proj_dir:
                logger.info(f"  [{sheet_name}] 精準掃描: {proj_dir}")
            else:
                logger.warning(f"  [{sheet_name}] 找不到專案根目錄，改用全域掃描")
                proj_dir = SEARCH_ROOT_DIR

            folder_idx, file_idx = build_path_index(proj_dir)
            n = fill_excel_paths(ws, folder_idx, file_idx, sheet_name)
            total_updated += n

        logger.info(f"路徑填寫完成，共更新 {total_updated} 筆")
        rebuild_index_sheet(wb)

        # ── 第二步：確保 NG_Reasons 分頁，讀取已有記錄 ──────────
        ws_ng = ensure_ng_sheet(wb)
        existing_keys, all_ng_rows = read_existing_ng(ws_ng)

        # ── 第三步：讀取韌體條目，比對步驟 ─────────────────────
        entries = scan_firmware_entries(wb)
        logger.info(f"找到 {len(entries)} 筆韌體條目")

        for entry in entries:
            folder_path  = entry["folder_path"]
            sheet_name   = entry["excel_sheet"]
            version_date = entry["version_date"]

            if not folder_path or not os.path.exists(folder_path):
                logger.debug(f"  [{sheet_name} {version_date}] 路徑不存在，跳過: {folder_path}")
                continue

            steps = parse_steps_from_define(folder_path)
            if not steps:
                logger.debug(f"  [{sheet_name} {version_date}] 未找到步驟定義")
                continue

            for step_code, step_name in steps.items():
                key = (sheet_name, version_date, step_code)
                if key in existing_keys:
                    continue

                # 版本繼承比對：ver_from="" 或 ver_from <= version_date 視為已覆蓋
                prev = find_prev_ng(all_ng_rows, sheet_name, step_code, version_date)
                if prev is not None:
                    existing_keys.add(key)
                    continue

                next_row = ws_ng.max_row + 1
                ws_ng.cell(row=next_row, column=COL_PRODUCT).value   = sheet_name
                ws_ng.cell(row=next_row, column=COL_VER_FROM).value  = version_date
                ws_ng.cell(row=next_row, column=COL_STEP_CODE).value = (
                    int(step_code) if step_code.isdigit() else step_code
                )
                ws_ng.cell(row=next_row, column=COL_STEP_NAME).value = step_name

                existing_keys.add(key)
                # 同時把新列加入 all_ng_rows，避免同一批次內重複繼承
                all_ng_rows.append({
                    "product":   sheet_name,
                    "ver_from":  version_date,
                    "step_code": step_code,
                    "desc":      "",
                    "err_data":  [],
                })
                pending_items.append({
                    "product":      sheet_name,
                    "version_date": version_date,
                    "step_code":    step_code,
                    "step_name":    step_name,
                    "inherited":    False,
                    "prev_ver":     "",
                })

        wb.save(backup_path)
        logger.info(f"副本已儲存: {os.path.basename(backup_path)}")

    except Exception as e:
        logger.error(f"執行時發生錯誤: {e}")
        import traceback
        logger.error(traceback.format_exc())
        input("按 Enter 鍵離開...")
        return
    finally:
        if wb:
            wb.close()
            del wb
            gc.collect()

    # ── 印出待補充清單 ─────────────────────────────────────────
    new_items       = [x for x in pending_items if not x["inherited"]]
    inherited_count = sum(1 for x in pending_items if x["inherited"])

    print("\n" + "=" * 60)
    if pending_items:
        print(f"  NG_Reasons 新增 {len(pending_items)} 筆（"
              f"全新 {len(new_items)} 筆 / 繼承前版自動填入 {inherited_count} 筆）")
        print()

        if new_items:
            print("  ★ 以下為全新步驟，請填寫不良描述與 ERR 欄位：")
            print(f"  {'產品':<16} {'版本起':<16} {'代碼':<6} 步驟名稱")
            print("  " + "-" * 56)
            for item in new_items:
                print(f"  {item['product']:<16} {item['version_date']:<16}"
                      f" {item['step_code']:<6} {item['step_name']}")
            print()
            print(f"  副本: {os.path.basename(backup_path)}")
            print()
            print("  ➜ 請開啟副本，填寫完成後執行 update_and_push.bat")
        else:
            print("  ✅ 全部為繼承前版，已自動填入，無需額外填寫")
            print()
            print("  ➜ 可直接執行 update_and_push.bat")
    else:
        print("  ✅ NG_Reasons 無需更新，所有步驟皆已有對應記錄")
        print()
        print("  ➜ 可直接執行 update_and_push.bat")
    print("=" * 60)

    input("\n按 Enter 鍵結束...")


if __name__ == "__main__":
    main()
