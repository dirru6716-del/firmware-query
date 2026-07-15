__version__ = "2026_0715_0814"
"""
scanner.py - 韌體資料庫自動掃描與更新工具
整合 update_records_20260319.py 全部功能
"""

import os
import re
import sys
import json
import shutil
import logging
import argparse
import gc
from datetime import datetime
from pathlib import Path
from glob import glob

import tkinter as tk
from tkinter import filedialog

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font

# ─── 設定區 ──────────────────────────────────────────────────────────────────
SEARCH_ROOT_DIR   = r"D:\yuanyu"          # 建立路徑索引的根目錄
SCAN_ROOT_DIRS    = [r"D:\yuanyu\OEM"]    # Fallback 韌體資料夾掃描根目錄
NETWORK_HEX_ROOT  = r"\\192.168.100.15\製造工程課\卓元裕\半測燒錄檔_output"
EXCEL_PATTERN     = "Software Change Log_*.xlsx"
OUTPUT_JSON       = "firmware_db.json"
EXCLUDE_SHEET_NAMES = {"Index", "索引製作方法", "NG_Reasons"}
MIN_VERSION_DATE  = "20200101_0000"       # 過舊版本排除門檻
# ─────────────────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - [%(levelname)s] - %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger()

VERSION_RE    = re.compile(r"_(\d{8}_\d{4})")
DEFINE_RE_MAP = {
    "result": re.compile(r"^#define\s+(\w+_result)\s+(\d+)",          re.MULTILINE),
    "data":   re.compile(r"^#define\s+(\w+_DATA(?:_\d+)?)\s+(\d+)",   re.MULTILINE | re.IGNORECASE),
    "step":   re.compile(r"^#define\s+(\w+_STEP(?:_\d+)?)\s+(\d+)",   re.MULTILINE),
}
EXCLUDE_KW = re.compile(r"last_data|last_time", re.IGNORECASE)

LAST_DATA_RE = re.compile(
    r"^#define\s+(TEST_result_last_data|TEST_time_last_data"
    r"|TEST_DATA_last_data|TEST_message_last_data)\s+(\d+)",
    re.MULTILINE | re.IGNORECASE,
)

# 在 .C / .H 原始碼中尋找版本字串靜態常數（支援 jig_ver、ver_data 等各種命名）
SOURCE_VER_RE = re.compile(
    r'unsigned\s+char\s+\w+\s*\[\s*\]\s*=\s*"'
    r'([^"]*(?:\d{9}_\d{4}|\d{8}_\d{4}|\d{4}_\d{4}_\d{4})[^"]*)"',
    re.IGNORECASE | re.MULTILINE,
)


# ═══════════════════════════════════════════════════════════════════════════════
# 第一階段：建立路徑索引（整合自 update_records_20260319.py）
# ═══════════════════════════════════════════════════════════════════════════════

def build_path_index(root_dir: str):
    """掃描 root_dir，建立路徑索引。
    folder_index: {資料夾名: [路徑, ...]}（保留所有候選，供後續依 sheet 名稱挑選）
    file_index:   {hex檔名: [所在資料夾, ...]}（同上）
    """
    logger.info(f"掃描目錄 {root_dir} 建立路徑索引（檔案較多請耐心等候）...")
    folder_index: dict[str, list[str]] = {}
    file_index:   dict[str, list[str]] = {}

    if not os.path.exists(root_dir):
        logger.error(f"路徑不存在: {root_dir}")
        return folder_index, file_index

    scan_count = 0
    for dirpath, dirnames, filenames in os.walk(root_dir):
        scan_count += 1
        if scan_count % 1000 == 0:
            logger.info(f"  持續掃描中... 已遍歷 {scan_count} 個子目錄")

        for d in dirnames:
            folder_index.setdefault(d, []).append(os.path.join(dirpath, d))
        for f in filenames:
            if f.lower().endswith(".hex"):
                # 無版本時間的 hex（開發中）直接忽略
                # 有版本時間但與所在資料夾路徑不一致的（複製帶入的舊 hex）也忽略
                m = VERSION_RE.search(f)
                if not m or m.group(1) not in dirpath:
                    continue
                file_index.setdefault(f, []).append(dirpath)

    logger.info(f"索引完成：資料夾 {len(folder_index)} 個，hex 檔 {len(file_index)} 個")
    return folder_index, file_index


def _best_path(candidates: list[str], sheet_name: str) -> str:
    """從多個候選路徑中，挑與 sheet_name 關鍵字最吻合的那條。"""
    if len(candidates) == 1:
        return candidates[0]
    # 把 sheet 名稱拆成關鍵字（英數字片段，忽略空白與特殊符號）
    keywords = re.findall(r"[A-Za-z0-9]+", sheet_name)
    def score(path: str) -> int:
        p = path.upper()
        return sum(1 for kw in keywords if kw.upper() in p)
    return max(candidates, key=score)


def get_real_cell(ws, cell):
    """取得合併儲存格的真實起始儲存格"""
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
    return cell


def fill_excel_paths(ws, folder_idx: dict, file_idx: dict, sheet_name: str):
    """將 F 欄路徑寫入工作表（含合併儲存格處理）。回傳更新筆數。"""
    max_row = ws.max_row
    f1_val = ws.cell(row=1, column=6).value

    if str(f1_val).strip() == "路徑":
        should_insert = False
        logger.info(f"  [{sheet_name}] F 欄已是路徑欄，直接覆寫")
    else:
        has_data = any(
            ws.cell(row=r, column=6).value for r in range(2, max_row + 1)
        )
        should_insert = has_data
        if has_data:
            logger.info(f"  [{sheet_name}] F 欄含有其他資料，插入新欄位")

    if should_insert:
        ws.insert_cols(6)

    ws.cell(row=1, column=6).value = "路徑"

    updated = 0
    for row in range(2, max_row + 1):
        c_cell = ws.cell(row=row, column=3)
        if isinstance(c_cell, MergedCell):
            continue

        key = str(c_cell.value).strip() if c_cell.value else ""
        f_raw  = ws.cell(row=row, column=6)
        f_real = get_real_cell(ws, f_raw)

        if not key:
            f_real.value = None
            continue

        if ".hex" in key.lower():
            candidates = file_idx.get(key, [])
        else:
            candidates = folder_idx.get(key, [])

        path = _best_path(candidates, sheet_name) if candidates else None
        f_real.value = path
        if path:
            updated += 1

    return updated


def rebuild_index_sheet(wb):
    """刪除舊 Index 分頁，重建超連結目錄"""
    if "Index" in wb.sheetnames:
        del wb["Index"]

    ws_idx = wb.create_sheet("Index", 0)
    ws_idx["A1"] = "工作表名稱"
    ws_idx["B1"] = "連結"

    for i, name in enumerate(wb.sheetnames):
        row = i + 2
        ws_idx.cell(row=row, column=1).value = name
        link_cell = ws_idx.cell(row=row, column=2)
        link_cell.value = "Go to Sheet"
        link_cell.hyperlink = f"#'{name}'!A1"
        try:
            link_cell.style = "Hyperlink"
        except Exception:
            link_cell.font = Font(color="0563C1", underline="single")

    ws_idx.column_dimensions["A"].width = 30
    ws_idx.column_dimensions["B"].width = 15
    logger.info("Index 分頁重建完成")


# ═══════════════════════════════════════════════════════════════════════════════
# 第二階段：從 Excel 讀取韌體條目
# ═══════════════════════════════════════════════════════════════════════════════

def read_excel_entries(wb) -> list[dict]:
    """讀取 Excel 各產品分頁，回傳韌體條目清單"""
    entries: list[dict] = []

    for sheet_name in wb.sheetnames:
        if sheet_name in EXCLUDE_SHEET_NAMES or sheet_name == "Index":
            continue

        ws = wb[sheet_name]
        for row in range(2, ws.max_row + 1):
            c_cell = ws.cell(row=row, column=3)
            if isinstance(c_cell, MergedCell) or not c_cell.value:
                continue

            c_val = str(c_cell.value).strip()
            m = VERSION_RE.search(c_val)
            if not m:
                continue

            version_date = m.group(1)
            if version_date.replace("_", "") < MIN_VERSION_DATE.replace("_", ""):
                continue
            if "notYet" in c_val:
                continue

            f_cell = ws.cell(row=row, column=6)
            f_real = get_real_cell(ws, f_cell)
            folder_path = str(f_real.value).strip() if f_real.value else ""

            a_val = ws.cell(row=row, column=1).value
            b_val = ws.cell(row=row, column=2).value
            d_val = ws.cell(row=row, column=4).value

            release_date = ""
            if d_val:
                if isinstance(d_val, datetime):
                    release_date = d_val.strftime("%Y-%m-%d")
                else:
                    release_date = str(d_val).strip()

            entries.append({
                "version_date": version_date,
                "folder_name":  c_val,
                "folder_path":  folder_path.replace("\\", "/"),
                "excel_sheet":  sheet_name.strip(),
                "case_name":    str(a_val).strip() if a_val else "",
                "change_notes": str(b_val).strip() if b_val else "",
                "release_date": release_date,
                # 以下欄位後續解析填入
                "test_results":  [],
                "test_data":     [],
                "test_steps":    {},
                "ng_reasons":    {},
                "result_count":  None,
                "time_count":    0,
                "data_count":    None,
                "message_count": None,
            })

    logger.info(f"從 Excel 讀取到 {len(entries)} 筆韌體條目")
    return entries


# ═══════════════════════════════════════════════════════════════════════════════
# 第三階段：解析 Define.h
# ═══════════════════════════════════════════════════════════════════════════════

def parse_define_h(folder_path: str) -> dict:
    """解析 Head_file/Define.h，回傳 {test_results, test_data, test_steps}"""
    define_path = os.path.join(folder_path, "Head_file", "Define.h")
    if not os.path.exists(define_path):
        logger.warning(f"  找不到 Define.h: {define_path}")
        return {}

    try:
        with open(define_path, "r", encoding="utf-8", errors="ignore") as f:
            content = f.read()
    except Exception as e:
        logger.warning(f"  無法讀取 Define.h: {e}")
        return {}

    result_map: dict[int, str] = {}
    data_map:   dict[int, str] = {}
    step_map:   dict[str, str] = {}

    for name, idx_str in DEFINE_RE_MAP["result"].findall(content):
        if EXCLUDE_KW.search(name):
            continue
        result_map[int(idx_str)] = name

    for name, idx_str in DEFINE_RE_MAP["data"].findall(content):
        if EXCLUDE_KW.search(name):
            continue
        data_map[int(idx_str)] = name

    for name, idx_str in DEFINE_RE_MAP["step"].findall(content):
        if EXCLUDE_KW.search(name):
            continue
        # step_map: 步驟數值字串 → 名稱（供查詢 TEST_STEP_DATA 時用）
        step_map[idx_str] = name

    last_data_map: dict[str, int] = {}
    for name, val in LAST_DATA_RE.findall(content):
        last_data_map[name.lower()] = int(val)

    test_results = [result_map[i] for i in sorted(result_map)]
    test_data    = [{"idx": i, "name": data_map[i]} for i in sorted(data_map)]

    return {
        "test_results":  test_results,
        "test_data":     test_data,
        "test_steps":    step_map,
        "result_count":  last_data_map.get("test_result_last_data"),
        "time_count":    last_data_map.get("test_time_last_data", 0),
        "data_count":    last_data_map.get("test_data_last_data"),
        "message_count": last_data_map.get("test_message_last_data"),
    }


def find_source_version(folder_path: str) -> str | None:
    """掃描 Source_file / Head_file 找版本靜態常數，回傳正規化後的版本字串。"""
    search_dirs = [
        os.path.join(folder_path, "Source_file"),
        os.path.join(folder_path, "Head_file"),
        folder_path,
    ]
    for d in search_dirs:
        if not os.path.isdir(d):
            continue
        for fname in os.listdir(d):
            if not fname.lower().endswith((".c", ".h")):
                continue
            fpath = os.path.join(d, fname)
            try:
                with open(fpath, "r", encoding="utf-8", errors="ignore") as f:
                    content = f.read()
            except Exception:
                continue
            m = SOURCE_VER_RE.search(content)
            if not m:
                continue
            raw = m.group(1).strip()
            digits = raw.replace("_", "")
            if len(digits) == 12:
                return digits[:8] + "_" + digits[8:]
            return raw  # 13 位（如 202600518_0818）原樣回傳，兩側一致即可
    return None


# ═══════════════════════════════════════════════════════════════════════════════
# 第四階段：讀取 NG_Reasons 分頁
# ═══════════════════════════════════════════════════════════════════════════════

def _parse_err_map(raw: str) -> dict[str, str]:
    """解析 '值=說明；分號分隔' 格式為 dict"""
    result: dict[str, str] = {}
    for part in raw.split(";"):
        part = part.strip()
        if "=" in part:
            k, _, v = part.partition("=")
            result[k.strip()] = v.strip()
    return result


def detect_err_columns(ws_ng) -> list[tuple[int, int]]:
    """
    掃描標題列，找出所有 (ERRn意義欄, ERRn對照欄) 對。
    欄位設計：從第 6 欄起，每 2 欄為一對：(意義, 對照)。
    回傳 [(意義欄號, 對照欄號), ...]，按 n 排序。
    """
    pairs: list[tuple[int, int]] = []
    max_col = ws_ng.max_column
    col = 6   # ERR 欄從第 6 欄開始
    while col + 1 <= max_col:
        h_mean = str(ws_ng.cell(row=1, column=col).value or "").strip()
        h_map  = str(ws_ng.cell(row=1, column=col + 1).value or "").strip()
        # 只要標頭含 ERR 且看起來是一對就收錄
        if re.search(r"ERR\d+", h_mean, re.IGNORECASE):
            pairs.append((col, col + 1))
        elif not h_mean and not h_map:
            break   # 遇到空標題就停止
        col += 2
    return pairs


def read_ng_reasons(wb) -> list[dict]:
    """
    讀取 NG_Reasons 分頁，支援動態 ERR 欄位數量。
    固定欄：A=產品, B=版本起, C=步驟代碼, D=步驟名稱, E=不良描述
    動態欄（從第 6 欄起，每對 2 欄）：ERR1意義, ERR1對照, ERR2意義, ERR2對照, ...
    """
    if "NG_Reasons" not in wb.sheetnames:
        logger.warning("NG_Reasons 分頁不存在，跳過")
        return []

    ws_ng    = wb["NG_Reasons"]
    err_cols = detect_err_columns(ws_ng)
    logger.info(f"NG_Reasons 偵測到 {len(err_cols)} 個 ERR 欄位對")

    rows: list[dict] = []
    for row in range(2, ws_ng.max_row + 1):
        product   = str(ws_ng.cell(row=row, column=1).value or "").strip()
        ver_from  = str(ws_ng.cell(row=row, column=2).value or "").strip()
        step_code = str(ws_ng.cell(row=row, column=3).value or "").strip()
        step_name = str(ws_ng.cell(row=row, column=4).value or "").strip()
        desc      = str(ws_ng.cell(row=row, column=5).value or "").strip()

        if not product or not step_code:
            continue

        # 動態讀取所有 ERR 欄對
        err_meanings: list[str] = []
        err_maps:     list[dict] = []
        for col_mean, col_map in err_cols:
            meaning = str(ws_ng.cell(row=row, column=col_mean).value or "").strip()
            map_raw = str(ws_ng.cell(row=row, column=col_map).value or "").strip()
            err_meanings.append(meaning)
            err_maps.append(_parse_err_map(map_raw))

        rows.append({
            "product":      product,
            "ver_from":     ver_from,
            "step_code":    step_code,
            "step_name":    step_name,
            "desc":         desc,
            "err_meanings": err_meanings,   # [ERR1意義, ERR2意義, ...]
            "err_maps":     err_maps,       # [{值:說明}, {值:說明}, ...]
        })

    logger.info(f"NG_Reasons 讀取 {len(rows)} 筆")
    return rows


def build_ng_reasons_for_entry(entry: dict, ng_raw: list[dict]) -> dict:
    """
    版本繼承邏輯：
    對每個步驟代碼，找所有符合 (product=sheet, step_code 相同, ver_from <= version_date) 的列，
    取 ver_from 最大（最近）的那筆。
    """
    sheet        = entry["excel_sheet"]
    version_date = entry["version_date"]
    ng_reasons: dict[str, dict] = {}

    candidates = [r for r in ng_raw if r["product"] == sheet]
    step_groups: dict[str, list[dict]] = {}
    for r in candidates:
        step_groups.setdefault(r["step_code"], []).append(r)

    for step_code, rows in step_groups.items():
        eligible = [
            r for r in rows
            if not r["ver_from"] or r["ver_from"] <= version_date
        ]
        if not eligible:
            continue
        best = max(eligible, key=lambda r: r["ver_from"])

        # 將動態 ERR 清單轉為 JSON 友善格式
        err_list: list[dict] = []
        for i, (meaning, err_map) in enumerate(
            zip(best["err_meanings"], best["err_maps"]), start=1
        ):
            if meaning or err_map:
                err_list.append({
                    "n":       i,
                    "meaning": meaning,
                    "map":     err_map,
                })

        ng_reasons[step_code] = {
            "step_name":   best["step_name"],
            "description": best["desc"],
            "err_data":    err_list,
        }

    return ng_reasons


# ═══════════════════════════════════════════════════════════════════════════════
# 第五階段：Fallback 掃描（補收 Excel 未收錄的資料夾）
# ═══════════════════════════════════════════════════════════════════════════════

def fallback_scan(known_versions: set[str]) -> list[dict]:
    """掃描 SCAN_ROOT_DIRS，補收 Excel 未收錄的韌體版本"""
    extra: list[dict] = []
    for root in SCAN_ROOT_DIRS:
        if not os.path.exists(root):
            logger.warning(f"Fallback 掃描目錄不存在: {root}")
            continue
        for dirpath, dirnames, _ in os.walk(root):
            for d in dirnames:
                if "notYet" in d:
                    continue
                m = VERSION_RE.search(d)
                if not m:
                    continue
                version_date = m.group(1)
                if version_date.replace("_", "") < MIN_VERSION_DATE.replace("_", ""):
                    continue
                if version_date in known_versions:
                    continue
                full_path = os.path.join(dirpath, d)
                extra.append({
                    "version_date": version_date,
                    "folder_name":  d,
                    "folder_path":  full_path.replace("\\", "/"),
                    "excel_sheet":  "",
                    "case_name":    "",
                    "change_notes": "",
                    "release_date": "",
                    "test_results":  [],
                    "test_data":     [],
                    "test_steps":    {},
                    "ng_reasons":    {},
                    "result_count":  None,
                    "time_count":    0,
                    "data_count":    None,
                    "message_count": None,
                })
                known_versions.add(version_date)

    logger.info(f"Fallback 掃描補收 {len(extra)} 筆")
    return extra


# ═══════════════════════════════════════════════════════════════════════════════
# 第六階段：複製 hex 檔到網路磁碟機
# ═══════════════════════════════════════════════════════════════════════════════

def copy_hex_to_network(entries: list[dict]):
    """將每個版本的 hex 檔複製到網路磁碟機，已存在者跳過"""
    if not os.path.exists(NETWORK_HEX_ROOT):
        logger.warning(f"網路磁碟機不可存取，跳過複製: {NETWORK_HEX_ROOT}")
        return

    copied = 0
    skipped = 0
    for entry in entries:
        folder_path = entry["folder_path"].replace("/", "\\")
        sheet_name  = entry["excel_sheet"]
        if not folder_path or not os.path.exists(folder_path) or not sheet_name:
            continue

        dest_dir = os.path.join(NETWORK_HEX_ROOT, sheet_name)
        try:
            os.makedirs(dest_dir, exist_ok=True)
        except Exception as e:
            logger.warning(f"  無法建立目錄 {dest_dir}: {e}")
            continue

        for fname in os.listdir(folder_path):
            if not fname.lower().endswith(".hex"):
                continue
            src  = os.path.join(folder_path, fname)
            dest = os.path.join(dest_dir, fname)
            if os.path.exists(dest):
                skipped += 1
                continue
            try:
                shutil.copy2(src, dest)
                logger.info(f"  複製: {fname} → {dest_dir}")
                copied += 1
            except Exception as e:
                logger.warning(f"  複製失敗 {fname}: {e}")

    logger.info(f"hex 複製完成：新增 {copied} 個，已存在跳過 {skipped} 個")


def copy_json_to_network(json_path: str):
    """將 firmware_db.json 複製到網路磁碟機根目錄（覆蓋舊檔）"""
    if not os.path.exists(NETWORK_HEX_ROOT):
        logger.warning(f"網路磁碟機不可存取，跳過 JSON 複製: {NETWORK_HEX_ROOT}")
        return

    dest = os.path.join(NETWORK_HEX_ROOT, os.path.basename(json_path))
    try:
        shutil.copy2(json_path, dest)
        logger.info(f"firmware_db.json 已複製到網路磁碟機: {dest}")
    except Exception as e:
        logger.warning(f"firmware_db.json 複製失敗: {e}")


def copy_excel_to_network(excel_path: str):
    """將 Excel 原始檔複製到網路磁碟機根目錄，檔名加上當下時間戳記與 _AI 後綴"""
    if not os.path.exists(NETWORK_HEX_ROOT):
        logger.warning(f"網路磁碟機不可存取，跳過 Excel 複製: {NETWORK_HEX_ROOT}")
        return

    stem, ext = os.path.splitext(os.path.basename(excel_path))
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    # 把原始檔名中已有的 _YYYYMMDD_HHMM 替換為當下時間，再加 _AI
    new_stem = re.sub(r"_\d{8}_\d{4}(_AI)?$", "", stem) + f"_{ts}_AI"
    new_fname = new_stem + ext
    dest = os.path.join(NETWORK_HEX_ROOT, new_fname)
    try:
        shutil.copy2(excel_path, dest)
        logger.info(f"Excel 已複製到網路磁碟機: {new_fname}")
    except Exception as e:
        logger.warning(f"Excel 複製失敗: {e}")


# ═══════════════════════════════════════════════════════════════════════════════
# 主流程
# ═══════════════════════════════════════════════════════════════════════════════

def select_excel() -> str | None:
    """開啟檔案對話框，讓使用者選擇 Excel 檔案"""
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    path = filedialog.askopenfilename(
        title="請選擇 Software Change Log Excel 檔案",
        filetypes=[("Excel 檔案", "*.xlsx *.xlsm"), ("所有檔案", "*.*")],
    )
    root.destroy()
    if not path:
        logger.warning("使用者取消選擇檔案")
        return None
    logger.info(f"已選擇: {os.path.basename(path)}")
    return path


def find_latest_excel(script_dir: str) -> str | None:
    """自動挑選 script_dir 中符合 EXCEL_PATTERN、修改時間最新的 Excel。
    排除 make_backup() 產生的備份副本（檔名含 _副本_），避免誤選備份造成連鎖再處理。
    """
    candidates = [
        p for p in glob(os.path.join(script_dir, EXCEL_PATTERN))
        if "_副本_" not in os.path.basename(p)
    ]
    if not candidates:
        logger.error(f"找不到符合 {EXCEL_PATTERN} 的 Excel（已排除備份副本）")
        return None
    latest = max(candidates, key=os.path.getmtime)
    logger.info(f"自動選用最新檔案: {os.path.basename(latest)}")
    return latest


def make_backup(excel_path: str) -> str | None:
    """建立時間戳記副本，回傳副本路徑"""
    fdir  = os.path.dirname(excel_path)
    fname = os.path.basename(excel_path)
    stem, ext = os.path.splitext(fname)
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = os.path.join(fdir, f"{stem}_副本_{ts}{ext}")
    try:
        shutil.copy2(excel_path, dest)
        logger.info(f"已建立副本: {os.path.basename(dest)}")
        return dest
    except Exception as e:
        logger.error(f"建立副本失敗: {e}")
        return None


def main(args):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    no_input = args.no_input

    def abort(msg: str):
        """失敗中止：無互動模式印錯誤並 exit(1)，互動模式維持原本 input 等待。"""
        if no_input:
            logger.error(msg)
            sys.exit(1)
        input(msg + " 按 Enter 鍵離開...")
        sys.exit(0)

    print("=" * 60)
    print(f"  韌體資料庫掃描工具 v{__version__}")
    print(f"  索引根目錄: {SEARCH_ROOT_DIR}")
    print(f"  網路磁碟機: {NETWORK_HEX_ROOT}")
    print("=" * 60)

    # ── 選擇 Excel ────────────────────────────────────────────
    if args.excel:
        excel_path = args.excel
        if not os.path.exists(excel_path):
            abort(f"指定的 Excel 不存在: {excel_path}")
        logger.info(f"使用指定檔案: {os.path.basename(excel_path)}")
    elif args.latest:
        excel_path = find_latest_excel(script_dir)
    else:
        excel_path = select_excel()

    if not excel_path:
        abort("未選擇檔案，程式中止。")

    # ── 建立副本（原檔完全不動）──────────────────────────────
    backup_path = make_backup(excel_path)
    if not backup_path:
        abort("備份失敗，程式中止。")

    # ── 第一階段：建立路徑索引 ────────────────────────────────
    folder_idx, file_idx = build_path_index(SEARCH_ROOT_DIR)

    # ── 第一階段（續）：填 Excel F 欄 ────────────────────────
    wb = None
    try:
        is_xlsm = backup_path.lower().endswith(".xlsm")
        wb = openpyxl.load_workbook(backup_path, keep_vba=is_xlsm)

        total_updated = 0
        for sheet_name in wb.sheetnames:
            if sheet_name in EXCLUDE_SHEET_NAMES or sheet_name == "Index":
                continue
            ws = wb[sheet_name]
            n  = fill_excel_paths(ws, folder_idx, file_idx, sheet_name)
            total_updated += n

        logger.info(f"路徑填寫完成，共更新 {total_updated} 筆")
        rebuild_index_sheet(wb)

        # ── 第二階段：讀取 Excel 韌體條目 ─────────────────────
        entries  = read_excel_entries(wb)
        ng_raw   = read_ng_reasons(wb)

        wb.save(backup_path)
        logger.info(f"副本已儲存: {os.path.basename(backup_path)}")

    except Exception as e:
        logger.error(f"處理 Excel 時發生錯誤: {e}")
        import traceback
        logger.error(traceback.format_exc())
        abort("處理 Excel 失敗，程式中止。")
    finally:
        if wb:
            wb.close()
            del wb
            gc.collect()

    # ── 第三階段：解析 Define.h + NG_Reasons ─────────────────
    logger.info("開始解析各版本 Define.h ...")
    for entry in entries:
        folder = entry["folder_path"].replace("/", "\\")
        if not folder or not os.path.exists(folder):
            continue
        parsed = parse_define_h(folder)
        if parsed:
            entry["test_results"]  = parsed.get("test_results", [])
            entry["test_data"]     = parsed.get("test_data", [])
            entry["test_steps"]    = parsed.get("test_steps", {})
            entry["result_count"]  = parsed.get("result_count")
            entry["time_count"]    = parsed.get("time_count", 0)
            entry["data_count"]    = parsed.get("data_count")
            entry["message_count"] = parsed.get("message_count")

        # 從原始碼讀取韌體實際送出的版本字串（覆蓋資料夾名稱版本）
        src_ver = find_source_version(folder)
        if src_ver:
            entry["version_date"] = src_ver

        entry["ng_reasons"] = build_ng_reasons_for_entry(entry, ng_raw)

    # ── 第五階段：Fallback 掃描 ───────────────────────────────
    known_versions = {e["version_date"] for e in entries}
    extra_entries  = fallback_scan(known_versions)

    for entry in extra_entries:
        folder = entry["folder_path"].replace("/", "\\")
        if folder and os.path.exists(folder):
            parsed = parse_define_h(folder)
            if parsed:
                entry["test_results"]  = parsed.get("test_results", [])
                entry["test_data"]     = parsed.get("test_data", [])
                entry["test_steps"]    = parsed.get("test_steps", {})
                entry["result_count"]  = parsed.get("result_count")
                entry["time_count"]    = parsed.get("time_count", 0)
                entry["data_count"]    = parsed.get("data_count")
                entry["message_count"] = parsed.get("message_count")

            src_ver = find_source_version(folder)
            if src_ver:
                entry["version_date"] = src_ver

    all_entries = entries + extra_entries

    # ── 第四階段：輸出 firmware_db.json ──────────────────────
    db = {
        "generated": datetime.now().isoformat(timespec="seconds"),
        "version":   __version__,
        "firmware":  sorted(all_entries, key=lambda e: e["version_date"], reverse=True),
    }

    json_path = os.path.join(script_dir, OUTPUT_JSON)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(db, f, ensure_ascii=False, indent=2)
    logger.info(f"已輸出 {OUTPUT_JSON}（{len(all_entries)} 筆韌體版本）")

    # ── 第六階段：複製 hex + Excel + JSON 到網路磁碟機 ────────
    copy_hex_to_network(all_entries)
    copy_excel_to_network(excel_path)
    copy_json_to_network(json_path)

    # ── 提示未填寫的 NG_Reasons ───────────────────────────────
    empty_ng = [
        r for r in ng_raw if not r["desc"]
    ]
    if empty_ng:
        print("\n" + "=" * 60)
        print("  ⚠  以下 NG_Reasons 說明尚未填寫：")
        for r in empty_ng:
            print(f"  產品={r['product']}  版本起={r['ver_from'] or '全版本'}"
                  f"  步驟={r['step_code']} {r['step_name']}")
        print("  請填寫完畢後再執行 update_and_push.bat")
        print("=" * 60)

    print("\n" + "=" * 60)
    print("  所有作業完成！")
    print(f"  副本: {os.path.basename(backup_path)}")
    print(f"  JSON: {OUTPUT_JSON}")
    print("=" * 60)
    if no_input:
        sys.exit(0)
    input("按 Enter 鍵結束...")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="韌體資料庫掃描工具")
    parser.add_argument("--excel", help="直接指定 Excel 路徑，不跳視窗")
    parser.add_argument("--latest", action="store_true",
                        help="自動選用本資料夾中最新的 Software Change Log_*.xlsx")
    parser.add_argument("--no-input", action="store_true",
                        help="無互動模式，不執行任何 input()，以 exit code 回報成敗")
    main(parser.parse_args())
